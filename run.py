# 项目运行入口
# 负责启动 Playwright 自动化测试并生成精美报告

import pytest
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from data.test_cases import TEST_CASES

# ==========================================
# 1. Excel 样式美化逻辑 (Excel Stylist)
# ==========================================

def finalize_excel_report(path, data):
    """
    [专业化处理] 为 Excel 报告添加企业级样式：
    包含颜色编码 (Status Coding)、自动列宽、表头冻结等。
    """
    df = pd.DataFrame(data)
    df.to_excel(path, index=False)

    wb = load_workbook(path)
    ws = wb.active

    # 定义颜色填充与字体样式
    fill_fail = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # 淡红
    fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # 淡绿
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # 深蓝
    font_header = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # 格式化表头 (Header)
    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center

    # 遍历数据行 (Data Rows)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # 序号与状态列居中
        ws.cell(row=row_idx, column=1).alignment = align_center
        
        status_cell = ws.cell(row=row_idx, column=4) # 假设 Status 在第 4 列
        status_val = str(status_cell.value).upper()
        
        if "FAIL" in status_val:
            status_cell.fill = fill_fail
        elif "PASS" in status_val:
            status_cell.fill = fill_pass
        
        status_cell.alignment = align_center
        # 测试日期列 (第 5 列)
        ws.cell(row=row_idx, column=5).alignment = align_center

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 6

    wb.save(path)

# ==========================================
# 2. 自动化执行主引擎 (Execution Engine)
# ==========================================

def run_integration_test():
    # 路径初始化
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    execution_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    report_dir = "reports"
    allure_results = "allure-results"
    
    if not os.path.exists(report_dir): os.makedirs(report_dir)

    excel_path = f"{report_dir}/QA_Integration_Report_{timestamp}.xlsx"
    html_path = f"{report_dir}/Technical_Evidence_{timestamp}.html"

    print(f"[{execution_date}] 系统就绪，开始执行端到端一致性校验...")

    # Pytest 参数配置
    args = [
        "tests/test_hardware_to_web.py",
        "-v",
        "--headed",
        f"--html={html_path}",
        "--self-contained-html",
        f"--alluredir={allure_results}",
        "--clean-alluredir"
    ]

    # 执行测试并捕获状态码 (0=Pass, 1=Fail)
    exit_code = pytest.main(args)

    # ==========================================
    # 3. 报告汇总阶段 (Report Aggregation)
    # ==========================================
    
    print("\n📈 正在生成结构化 Excel 报告...")
    
    # 整合 Excel 数据
    summary_data = []
    for case in TEST_CASES:
        # 在真实场景中，这里的 Status 可以通过解析 allure-results JSON 获取
        # 演示逻辑：
        status_text = "PASSED" if exit_code == 0 else "FAILED / CHECK LOG"
        
        summary_data.append({
            "No. (序号)": case['id'],
            "Test Scenario (测试场景描述)": case['scenario'],
            "expected_input (期望输入)": case['expected_input'],
            "expected_output (期望结果)": case['expected_output'],
            "expected_min (期望最小值)": case.get('expected_min'),
            "expected_max (期望最大值)": case.get('expected_max'),
            "Monitoring Point (监控点位)": case['point'],
            "Test Status (测试状态)": status_text,
            "Execution Date (测试日期)": execution_date,
            "Hardware Data Type (数据类型)": case['type']
        })

    # 美化并保存 Excel
    finalize_excel_report(excel_path, summary_data)

    print("-" * 65)
    print(f"🎉 测试任务圆满结束！交付件清单：")
    print(f"▶ 管理层报表 (Excel): {os.path.abspath(excel_path)}")
    print(f"▶ 快速审计文档 (HTML): {os.path.abspath(html_path)}")
    print(f"▶ 全功能报告 (Allure): 运行 'allure serve {allure_results}'")
    print("-" * 65)

if __name__ == "__main__":
    run_integration_test()